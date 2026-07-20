"""
build-pillar-measures.py — Geoff's measures sheet → data/pillar-measures.json
=============================================================================
Reads:
  - data/source/StrategicPlan_measures.xlsx  (downloaded export of Geoff's
        Google Sheet "StrategicPlan_measures", sheet tab "Pillars"; pass an
        alternate path as the first CLI arg)
  - data/DIM_Measures.csv                    (site-side measure registry: names,
        labels, sort order, BestInNationGoal flag)
  - data/DIM_Pillars.csv                     (pillar names)
  - data/pillar-measures.json                (merge base; hand-authored fields
        preserved — safe to re-run on every data wave)

Produces:
  - data/pillar-measures.json                (same schema as measures.json, so
        pillar.html can reuse the Best-in-Nation chart component)

Usage:
  python data/build-pillar-measures.py [path/to/export.xlsx]

Requires openpyxl (already in the devcontainer; `pip install openpyxl` elsewhere).

Why this exists (plan: notes/plan-pillar-measures-20260717.md, approved 2026-07-17):
  Geoff's sheet is a live scratchpad — data arrives in waves as he finalizes
  measures with agency teams (flagged Y in the "Finalized?" column). Each wave
  is a re-run of this script, not a re-clean. The BiN pipeline
  (build-measures.py) is separate and untouched; per decision 1, BiN-flagged
  measures appear only on the best-in-nation page and are skipped here.

Key rules (decision numbers reference the plan doc):
  - Strict Y filter: the Finalized? column contains prose; only a trimmed,
    case-folded "y" counts.
  - Duplicate Measure IDs among Y rows → HARD ABORT (mid-edit sheet state;
    refuse to guess which row Geoff means).
  - Widened ID pattern P#.M#[a-z] for Geoff's sub-splits (decision 4).
    DIM_Measures.csv is the registry: a Y-flagged ID missing from DIM is
    EXCLUDED with a loud warning (Andy adds the DIM row, re-runs). Base and
    sub-ID both Y (e.g. P2.M2 + P2.M2a) → Geoff is mid-split; exclude the pair.
  - P5.M2 (NCSIS milestone, values all 1) excluded by name (decision 2) until
    Geoff resolves its chartability.
  - Values: percent cells in the export are FRACTIONS with a % number format
    (0.877 + "0.0%"); they are converted to percent points (87.7) to match the
    measures.json convention. "-", ".", "?", "TBD", blank → null, never 0.
  - Status (decision 3) is DERIVED, direction-aware (decrease measures invert),
    and regenerated every run; the hand-override lives in statusOverride
    {type, label}, which is never touched and wins in the renderer.
  - Source cells are often Geoff's meeting notes. Only a clean "prose + one
    URL" (or a bare URL) is auto-split into sourceLabel/sourceUrl; anything
    messier is left null and sent to the warning list for hand-authoring via
    sourceHtml (preserved field; the BiN renderer already prefers it). Raw
    scratchpad prose must not leak into this public repo's JSON.
  - Chief / NCDPI Business Owner / Responsible DPI Person columns are never
    read (decision 6 — staff names, public repo).
  - The end-of-run WARNING SUMMARY is the per-wave to-do list.
  - The gaps report (data/measure-gaps.md) also reconciles DIM_Measures.csv
    against ALL sheet rows both ways every run (punch #3), so unregistered
    sub-IDs get DIM rows before a wave flips them Y.

Author: Andy Baxter / Claude  |  2026-07-17
"""

import json
import csv
import datetime
import math
import os
import re
import sys

import openpyxl

# Windows console encoding — needed for em-dash output in labels.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, Exception):
    pass


# --- Configuration --------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

DEFAULT_XLSX = os.path.join(SCRIPT_DIR, "source", "StrategicPlan_measures.xlsx")
SHEET_NAME = "Pillars"
OUTPUT_JSON = os.path.join(SCRIPT_DIR, "pillar-measures.json")
DIM_MEASURES = os.path.join(SCRIPT_DIR, "DIM_Measures.csv")
DIM_PILLARS = os.path.join(SCRIPT_DIR, "DIM_Pillars.csv")

# Widened ID pattern — this pipeline only (decision 4). build-measures.py
# keeps its strict P#.M# regex, so DIM sub-ID rows are invisible to it.
MEASURE_ID_RE = re.compile(r"^P\d+\.M\d+[a-z]?$")

# Decision 2: excluded by name, with a warning, until resolved with Geoff.
EXCLUDED_IDS = {
    "P5.M2": "NCSIS milestone measure (values all 1) — not chartable in the "
             "two-chart format; raise with Geoff",
}

# Sheet columns we consume, by header text (matched after trim). Column ORDER
# in the sheet is not trusted — Geoff rearranges. Staff-name columns (Chief,
# NCDPI Business Owner, Responsible DPI Person) are deliberately absent here.
COL_FINALIZED = "Finalized?"
COL_PILLAR = "Pillar"
COL_ID = "Measure ID"
COL_MEASURE = "Measure"
COL_AVAILABLE = "When Available?"
COL_SOURCE = "Source"
COL_WHY = "WhyMeasureMatters"
COL_CONTEXT = "MeasureContextNote"
# Year columns are found by regex on the header ("Current (2024)",
# "2025 (Actual)", "2026 (Target)" ...) so added years don't break the build.
YEAR_HEADER_RE = re.compile(r"(20\d\d)")

# Fields regenerated every run vs. preserved from the existing JSON:
#   REGENERATED: everything sourced from the sheet or derived (dataSeries,
#     valueFormat, yAxisMax, statusType, statusLabel, currentValue, goal, ...)
#   PRESERVED (hand-authored, never overwritten): iconUrl, definition,
#     metricText, sourceHtml, moreDataUrl, moreDataText, badge* (unused for
#     pillar measures but kept for schema parity), currentDescription,
#     statusOverride, and notes/whyItCounts when the sheet cell is blank.
PRESERVE_NOTE = "see header comment"  # documentation anchor only


# --- Warning collector -----------------------------------------------------

WARNINGS = []


def warn(measure_id, message):
    """Collect a warning for the end-of-run summary (also printed inline)."""
    WARNINGS.append((measure_id or "—", message))
    print(f"  WARN [{measure_id or '—'}]: {message}")


def abort(message):
    print(f"\n  ABORT: {message}")
    print("  No output written.")
    sys.exit(1)


# --- Cell parsing ----------------------------------------------------------

NULL_TOKENS = {"", "-", ".", "?", "tbd", "n/a", "na"}


def clean_text(value):
    """Collapse whitespace; return '' for None/blank. Dates → 'Month YYYY'."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):  # datetime from a date-formatted cell
        return value.strftime("%B %Y")
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_cell(cell):
    """Parse one data cell → (value_in_display_units, kind) where kind is one
    of 'pct', 'dollar', 'comma', 'plain', or (None, None) for empty/null.

    Numeric percent cells hold FRACTIONS (0.877 + '%' number format) and are
    converted to percent points; string cells like '72.10%' are already in
    percent points and are not rescaled.
    """
    raw = cell.value
    fmt = cell.number_format or ""

    if raw is None:
        return None, None

    if isinstance(raw, (int, float)):
        if "%" in fmt:
            # round() kills float noise from the fraction→percent-points
            # conversion (0.55 * 100 = 55.00000000000001 in binary floats).
            return round(float(raw) * 100.0, 6), "pct"
        if "$" in fmt:
            return float(raw), "dollar"
        if "," in fmt:
            return float(raw), "comma"
        return float(raw), "plain"

    s = str(raw).strip()
    if s.lower() in NULL_TOKENS:
        return None, None
    kind = "plain"
    if s.endswith("%"):
        kind = "pct"
    elif s.startswith("$"):
        kind = "dollar"
    elif "," in s:
        kind = "comma"
    s = s.replace("$", "").replace("%", "").replace(",", "").strip()
    try:
        return float(s), kind
    except ValueError:
        return None, "unparseable"


def infer_value_format(kinds, values):
    """Map the parsed cell kinds to a measures.json valueFormat string.

    '$#,###' is NEW to this pipeline (BiN has no dollar measures) — chunk C
    must add a matching case to pillar.html's formatValue before a dollar
    measure ships.
    """
    if "pct" in kinds:
        return "#.#%"
    if "dollar" in kinds:
        return "$#,###"
    if "comma" in kinds:
        return "#,###"
    if any(v is not None and float(v) != int(v) for v in values):
        return "#.#"
    # Whole numbers: use thousands separators when the magnitude needs them.
    if any(v is not None and abs(v) >= 1000 for v in values):
        return "#,###"
    return "#"


def format_value(value, value_format):
    """Format a numeric value per its valueFormat string (BiN conventions)."""
    if value is None:
        return None
    if value_format == "#.#%":
        return f"{value:.1f}%"
    if value_format == "#.#":
        return f"{value:.1f}"
    if value_format == "#,###":
        return f"{int(round(value)):,}"
    if value_format == "$#,###":
        return f"${int(round(value)):,}"
    return f"{int(round(value))}"


def slugify(name):
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return re.sub(r"-+", "-", s).strip("-")


# --- Source splitter --------------------------------------------------------

URL_RE = re.compile(r"https?://\S+")


def split_source(measure_id, raw):
    """Split a Source cell into (sourceLabel, sourceUrl) — conservatively.

    Auto-split ONLY when the cell is a bare URL, or short prose plus exactly
    one URL with no scratchpad markers ('?', ';'). Anything else returns
    (None, None) and a hand-review warning: Geoff's source cells are often
    meeting notes, and raw scratchpad prose must not be published into this
    public repo's JSON. Hand-authored sourceHtml (preserved field) is the
    intended fix-up channel — the renderer prefers it.
    """
    text = clean_text(raw)
    if not text:
        return None, None

    urls = URL_RE.findall(text)
    # Trim separator punctuation only — never parens, which are part of the
    # prose ("... (Broaden Successful Participation in Advanced Courses)").
    prose = URL_RE.sub("", text).strip(" ;:,–—-").strip()

    if len(urls) == 1 and not prose:
        warn(measure_id, "Source is a bare URL — add a human-readable "
                         "sourceLabel by hand (or via sourceHtml)")
        return None, urls[0].rstrip(").,;")

    if len(urls) == 1 and prose and "?" not in prose and ";" not in prose:
        return prose, urls[0].rstrip(").,;")

    warn(measure_id, f"Source cell needs hand review (multi-part or "
                     f"scratchpad prose) — left null; hand-author sourceHtml. "
                     f"Raw: {text[:120]}")
    return None, None


# --- Readers ----------------------------------------------------------------

def read_dim_measures():
    """{MeasureID: {name, label, sort, bestInNation, goalFallback}} — widened regex."""
    out = {}
    with open(DIM_MEASURES, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            mid = (row.get("MeasureID") or "").strip()
            if not MEASURE_ID_RE.match(mid):
                continue
            bin_raw = (row.get("BestInNationGoal") or "").strip()
            try:
                sort = float(row.get("MeasureSort") or "")
            except ValueError:
                sort = None
            out[mid] = {
                "name": clean_text(row.get("MeasureName")),
                "label": clean_text(row.get("MeasureLbl")),
                "goalFallback": clean_text(row.get("MeasureGoal")),
                "sort": sort,
                "bestInNation": bool(bin_raw),
            }
    return out


def read_pillars():
    """{pillar number (int): PillarName}."""
    out = {}
    with open(DIM_PILLARS, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            pid = (row.get("PillarID") or "").strip()
            m = re.match(r"^P(\d+)$", pid)
            if m and clean_text(row.get("PillarName")):
                out[int(m.group(1))] = clean_text(row.get("PillarName"))
    return out


def load_existing():
    """Existing pillar-measures.json keyed by measureId (merge base)."""
    if not os.path.exists(OUTPUT_JSON):
        print(f"  Note: {OUTPUT_JSON} doesn't exist yet; starting fresh.")
        return {}
    try:
        with open(OUTPUT_JSON, encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        abort(f"could not parse existing {OUTPUT_JSON}: {e} — fix or remove "
              f"it before re-running (refusing to clobber hand-authored fields).")
    return {m["measureId"]: m for m in data if m.get("measureId")}


def read_sheet(xlsx_path):
    """Read the Pillars tab → (Y-flagged row dicts, ID census).

    Y rows keep raw cells for format-aware parsing; duplicate Y-row Measure
    IDs abort. The census records EVERY non-blank Measure ID (row number +
    Finalized? text) so the gaps report can reconcile DIM against the whole
    sheet — not-yet-Y rows are exactly the ones worth registering in DIM
    before a wave flips them Y (punch list 2026-07-20 #3)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        abort(f"sheet tab '{SHEET_NAME}' not found in {xlsx_path} "
              f"(tabs: {wb.sheetnames})")
    ws = wb[SHEET_NAME]

    headers = {}
    for idx, cell in enumerate(ws[1]):
        h = clean_text(cell.value)
        if h:
            headers[h] = idx
    for required in (COL_FINALIZED, COL_ID, COL_MEASURE):
        if required not in headers:
            abort(f"required column '{required}' not found in sheet header "
                  f"row (headers: {sorted(headers)})")

    # Year columns: header regex, classified baseline (Current/Actual) vs Target.
    year_cols = {}  # year int -> (col idx, 'baseline'|'target')
    for h, idx in headers.items():
        m = YEAR_HEADER_RE.search(h)
        if not m:
            continue
        year = int(m.group(1))
        role = "target" if "target" in h.lower() else "baseline"
        year_cols[year] = (idx, role)
    if not year_cols:
        abort("no year columns recognized in the sheet header row")

    rows = []
    census = []  # (row_num, measure_id, finalized_text) for ALL rows with an ID
    seen_ids = {}
    for row in ws.iter_rows(min_row=2):
        finalized = clean_text(row[headers[COL_FINALIZED]].value)
        mid = clean_text(row[headers[COL_ID]].value)
        if mid:
            census.append((row[0].row, mid, finalized))
        if finalized.casefold() != "y":
            continue
        if mid in seen_ids:
            abort(f"duplicate Measure ID '{mid}' among Y-flagged rows "
                  f"(rows {seen_ids[mid]} and {row[0].row}) — Geoff's sheet is "
                  f"mid-edit; resolve the duplicate and re-run.")
        seen_ids[mid] = row[0].row

        def cell_of(col_name):
            return row[headers[col_name]] if col_name in headers else None

        rows.append({
            "row_num": row[0].row,
            "measureId": mid,
            "pillar_raw": cell_of(COL_PILLAR).value if COL_PILLAR in headers else None,
            "measure_text": clean_text(cell_of(COL_MEASURE).value),
            "available": clean_text(cell_of(COL_AVAILABLE).value) if COL_AVAILABLE in headers else "",
            "source_raw": cell_of(COL_SOURCE).value if COL_SOURCE in headers else None,
            "why": clean_text(cell_of(COL_WHY).value) if COL_WHY in headers else "",
            "context": clean_text(cell_of(COL_CONTEXT).value) if COL_CONTEXT in headers else "",
            "year_cells": {yr: row[idx] for yr, (idx, _) in year_cols.items()},
            "year_roles": {yr: role for yr, (idx, role) in year_cols.items()},
        })
    return rows, census


# --- Derivations ------------------------------------------------------------

def build_data_series(row):
    """Sheet year cells → measures.json dataSeries (+ inferred valueFormat).

    Convention (matches measures.json): observed actuals go in `baseline`,
    targets in `target`, `actual` reserved for future post-baseline tracking.
    All year rows are kept even when null — chart x-axis spacing depends on
    consistent years (build-measures.py convention).
    """
    series = []
    kinds = []
    values = []
    for year in sorted(row["year_cells"]):
        cell = row["year_cells"][year]
        val, kind = parse_cell(cell)
        if kind == "unparseable":
            warn(row["measureId"],
                 f"unparseable value {cell.value!r} in {year} column — treated as null")
            val, kind = None, None
        if kind:
            kinds.append(kind)
        if val is not None:
            values.append(val)
        role = row["year_roles"][year]
        series.append({
            "year": str(year),
            "baseline": val if role == "baseline" else None,
            "target": val if role == "target" else None,
            "actual": None,
        })
    if len(set(kinds)) > 1:
        warn(row["measureId"],
             f"mixed cell formats across year columns ({sorted(set(kinds))}) — "
             f"check the sheet row")
    value_format = infer_value_format(kinds, values) if values else None
    return series, value_format, values


# "Nice" axis-max mantissas (punch list 2026-07-20 #4a). Andy's hand-set
# BiN maxes follow the same 1.2×-then-round shape (ACT 20 → 24, AP perf
# 75 → 90); this rounds the derived value UP to the nearest nice mantissa
# so pillar charts read the same (68.4 → 70, 3,720 → 4,000).
NICE_MANTISSAS = [1, 1.2, 1.5, 2, 2.5, 3, 3.5, 4, 5, 6, 7, 8, 9, 10]


def derive_y_axis_max(values, value_format):
    """BiN convention: 1.2 × series max, rounded up to a nice number,
    capped at 100 for percent measures."""
    if not values:
        return None
    y_max = max(values) * 1.2
    if y_max > 0:
        exp = math.floor(math.log10(y_max))
        mantissa = y_max / 10 ** exp
        for nice in NICE_MANTISSAS:
            if nice >= mantissa - 1e-9:
                y_max = nice * 10 ** exp
                break
    if value_format == "#.#%":
        y_max = min(y_max, 100.0)
    return round(float(y_max), 1)


def derive_status(measure_id, series, value_format):
    """Direction-aware derived status (decision 3). Regenerated every run;
    hand overrides belong in statusOverride, which this never touches.

    direction = sign(final target − earliest observed value); decrease
    measures (e.g. P4.M4 chronic absenteeism) invert "improvement".
      - latest actual meets/beats the next upcoming target → on-target
      - improved vs. the prior actual → approaching-target
      - single observed value → baseline ("Baseline Year — X")
      - regressed vs. prior actual → null + warning (hand-author; the script
        will not print "Approaching target" over a decline)
    """
    actuals = [(int(e["year"]), e["baseline"]) for e in series if e["baseline"] is not None]
    targets = [(int(e["year"]), e["target"]) for e in series if e["target"] is not None]
    if not actuals:
        return None, None

    latest_year, latest = actuals[-1]
    label_value = format_value(latest, value_format)

    if len(actuals) == 1:
        return "baseline", f"Baseline Year — {label_value}"

    if not targets:
        warn(measure_id, "no targets in sheet row — status not derived")
        return None, None

    final_target = targets[-1][1]
    earliest = actuals[0][1]
    decrease = final_target < earliest

    def improved(new, old):
        return new < old if decrease else new > old

    def meets(actual, target):
        return actual <= target if decrease else actual >= target

    # Target for the latest actual year if present; else the next upcoming one.
    same_or_next = [t for (y, t) in targets if y >= latest_year]
    comparison_target = same_or_next[0] if same_or_next else final_target

    if meets(latest, comparison_target):
        return "on-target", f"On Target — {label_value}"
    if improved(latest, actuals[-2][1]):
        return "approaching-target", f"Approaching Target — {label_value}"

    warn(measure_id,
         f"latest actual ({label_value}) regressed vs. prior year — status "
         f"left null; hand-author via statusOverride if needed")
    return None, None


# --- Assembly ---------------------------------------------------------------

def build_measure(row, dim_entry, pillar_names, existing_map):
    mid = row["measureId"]
    existing = existing_map.get(mid, {})

    pillar_num = int(mid.split(".")[0][1:])

    # Cross-check the sheet's Pillar column against the ID prefix.
    pillar_raw = row["pillar_raw"]
    if pillar_raw is not None:
        try:
            if int(float(pillar_raw)) != pillar_num:
                warn(mid, f"sheet Pillar column ({pillar_raw}) != ID prefix P{pillar_num}")
        except (ValueError, TypeError):
            warn(mid, f"sheet Pillar column not numeric: {pillar_raw!r}")

    series, value_format, values = build_data_series(row)
    status_type, status_label = derive_status(mid, series, value_format)
    source_label, source_url = split_source(mid, row["source_raw"])

    name = dim_entry["name"] or mid
    latest_actual = next(
        (e["baseline"] for e in reversed(series) if e["baseline"] is not None), None
    )

    # Preserve-on-blank: sheet cell wins when non-empty, else keep existing.
    def sheet_or_existing(sheet_val, key):
        return sheet_val if sheet_val else existing.get(key)

    return {
        "id": slugify(name),
        "measureId": mid,
        "name": name,
        "pillarNumber": pillar_num,
        "pillarName": pillar_names.get(pillar_num, ""),
        "menuLabel": dim_entry["label"] or name,
        "menuSubLabel": existing.get("menuSubLabel"),
        "menuTooltip": existing.get("menuTooltip", ""),
        "iconUrl": existing.get("iconUrl"),                    # PRESERVE
        "goal": row["measure_text"] or dim_entry["goalFallback"],
        "definition": existing.get("definition"),              # PRESERVE
        "metricText": existing.get("metricText"),              # PRESERVE
        "valueFormat": value_format,
        "yAxisMax": derive_y_axis_max(values, value_format),
        "sourceLabel": sheet_or_existing(source_label, "sourceLabel"),
        "sourceUrl": sheet_or_existing(source_url, "sourceUrl"),
        "sourceHtml": existing.get("sourceHtml"),              # PRESERVE
        "moreDataUrl": existing.get("moreDataUrl"),            # PRESERVE
        "moreDataText": existing.get("moreDataText"),          # PRESERVE
        "badgeImageUrl": existing.get("badgeImageUrl"),        # PRESERVE (schema parity)
        "badgeAltText": existing.get("badgeAltText"),          # PRESERVE
        "badgePressReleaseUrl": existing.get("badgePressReleaseUrl"),  # PRESERVE
        "statusType": status_type,      # derived, regenerated every run
        "statusLabel": status_label,    # derived, regenerated every run
        "statusOverride": existing.get("statusOverride"),      # PRESERVE — hand
                                        # {type, label}; renderer prefers it
        "currentValue": format_value(latest_actual, value_format),
        "currentDescription": existing.get("currentDescription"),  # PRESERVE
        "whyItCounts": sheet_or_existing(row["why"], "whyItCounts"),
        "nextUpdate": sheet_or_existing(row["available"], "nextUpdate"),
        "notes": sheet_or_existing(row["context"], "notes"),
        "dataSeries": series,
    }


def sort_key(measure, dim_lookup):
    """Pillar number, then DIM MeasureSort, then M-number (+ sub-letter)."""
    mid = measure["measureId"]
    dim_sort = dim_lookup.get(mid, {}).get("sort")
    m = re.match(r"^P\d+\.M(\d+)([a-z]?)$", mid)
    m_num = int(m.group(1)) + (ord(m.group(2)) - ord("a") + 1) / 100 if m.group(2) else int(m.group(1))
    return (measure["pillarNumber"], dim_sort if dim_sort is not None else m_num, mid)


# --- Gaps report (punch list 2026-07-20 #2) ---------------------------------

GAPS_MD = os.path.join(SCRIPT_DIR, "measure-gaps.md")


def dim_reconciliation(dim_lookup, census):
    """Two-way DIM_Measures.csv <-> sheet audit (punch list 2026-07-20 #3).

    Compares against ALL sheet rows, not just Y ones: an unregistered ID is
    harmless until a wave flips it Y and the pipeline excludes it — this
    section exists so DIM rows get added BEFORE that happens, and so DIM
    rows the sheet restructured away (e.g. a base ID split into sub-IDs)
    can't linger. IDs and row numbers only — no sheet prose (public repo)."""
    findings = []
    sheet_ids = {}
    for row_num, mid, _fin in census:
        sheet_ids.setdefault(mid, []).append(row_num)

    for mid, row_nums in sorted(sheet_ids.items()):
        if len(row_nums) > 1:
            findings.append(
                f"sheet ID `{mid}` appears on {len(row_nums)} rows "
                f"({', '.join(map(str, row_nums))}) — needs distinct sub-IDs: "
                f"2+ of them Y aborts the build; exactly one Y would chart "
                f"under the ambiguous shared ID")
        if not MEASURE_ID_RE.match(mid):
            findings.append(
                f"sheet row {row_nums[0]}: Measure ID `{mid}` doesn't match "
                f"P#.M#[a-z] — needs a real ID before it can chart")
        elif mid not in dim_lookup:
            findings.append(
                f"sheet ID `{mid}` is not in DIM_Measures.csv — it will be "
                f"EXCLUDED if flagged Y; add the DIM row (name + sort) now")

    for mid in sorted(dim_lookup):
        if mid not in sheet_ids:
            findings.append(
                f"DIM ID `{mid}` is no longer anywhere in the sheet — "
                f"restructured out? Remove or remap the DIM row")
    return findings


def measure_gaps(row, measure):
    """Missing-field notes for one charted measure. Sheet-column gaps are
    phrased for pinging Geoff; hand-authored gaps are Andy's side."""
    out = []
    if not row["available"] and not measure["nextUpdate"]:
        out.append('sheet "When Available?" — drives the card\'s **Next update** line')
    if not row["why"] and not measure["whyItCounts"]:
        out.append('sheet "WhyMeasureMatters" — the **Why it counts** blurb')
    if not row["context"] and not measure["notes"]:
        out.append('sheet "MeasureContextNote" — extra context under the chart')
    # Renderer needs sourceHtml, or at least a sourceLabel (a bare
    # sourceUrl with no label renders no Source line at all).
    if not (measure["sourceHtml"] or measure["sourceLabel"]):
        out.append("no usable Source — card renders **no Source line** "
                   "(hand-author sourceHtml, or get a clean source cell from Geoff)")
    actuals = [e for e in measure["dataSeries"] if e["baseline"] is not None]
    targets = [e for e in measure["dataSeries"] if e["target"] is not None]
    if len(actuals) < 2:
        out.append("only one year of actuals so far")
    if not targets:
        out.append("no targets in the sheet row")
    return out


def write_gaps_report(gaps, warnings, dim_findings):
    """data/measure-gaps.md — tracked, regenerated every run. Andy pastes
    from it to ping Geoff after each wave.

    Deliberately NEVER quotes raw sheet cell contents: the repo is public
    (GitHub Pages) and Geoff's scratchpad prose stays out of it, same
    reason data/source/ is gitignored. Console warnings keep the raw
    snippets; here they are stripped."""
    today = datetime.date.today().isoformat()
    lines = [
        "# Measure gaps — per-wave to-do list",
        "",
        f"_Generated by build-pillar-measures.py on {today}; regenerated every "
        "run — do not hand-edit. Raw sheet text is deliberately not quoted "
        "here (public repo)._",
        "",
        "## Missing fields by charted measure",
        "",
    ]
    any_gap = False
    for mid in sorted(gaps):
        if not gaps[mid]:
            continue
        any_gap = True
        lines.append(f"### {mid}")
        lines += [f"- {g}" for g in gaps[mid]]
        lines.append("")
    if not any_gap:
        lines += ["All charted measures have their display fields. 🎉", ""]

    lines += ["## DIM ↔ sheet reconciliation (all rows, not just finalized)", ""]
    if dim_findings:
        lines += [f"- {f}" for f in dim_findings]
    else:
        lines.append("- in sync — every sheet ID is registered in "
                     "DIM_Measures.csv and every DIM ID is still in the sheet")
    lines.append("")

    lines += ["## Pipeline warnings this run", ""]
    if warnings:
        for mid, msg in warnings:
            lines.append(f"- **{mid}**: {msg.split(' Raw:')[0].strip()}")
    else:
        lines.append("- none")
    lines.append("")

    with open(GAPS_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    n = sum(1 for g in gaps.values() if g)
    print(f"  Gaps report: {GAPS_MD} ({n} of {len(gaps)} measures have missing fields)")


# --- Main -------------------------------------------------------------------

def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx_path):
        abort(f"input not found: {xlsx_path}")

    print(f"Reading {xlsx_path}")
    dim_lookup = read_dim_measures()
    pillar_names = read_pillars()
    existing_map = load_existing()
    y_rows, id_census = read_sheet(xlsx_path)

    print(f"  - DIM_Measures.csv: {len(dim_lookup)} registry rows")
    print(f"  - DIM_Pillars.csv: {len(pillar_names)} pillars")
    print(f"  - existing pillar-measures.json: {len(existing_map)} measures")
    print(f"  - sheet: {len(y_rows)} Y-flagged rows")

    y_ids = {r["measureId"] for r in y_rows}
    charted = []
    skipped_bin = []
    gaps = {}  # measureId -> list of missing-field notes (gaps report, punch list #2)

    for row in y_rows:
        mid = row["measureId"]

        if not MEASURE_ID_RE.match(mid):
            warn(mid, f"row {row['row_num']}: Measure ID doesn't match "
                      f"P#.M#[a-z] — EXCLUDED")
            continue

        if mid in EXCLUDED_IDS:
            warn(mid, f"EXCLUDED — {EXCLUDED_IDS[mid]}")
            continue

        dim_entry = dim_lookup.get(mid)
        if dim_entry is None:
            warn(mid, "Y-flagged but not in DIM_Measures.csv — EXCLUDED. "
                      "Add the DIM row (name + sort) and re-run.")
            continue

        if dim_entry["bestInNation"]:
            skipped_bin.append(mid)  # decision 1: BiN page only
            continue

        # Base + sub-ID both Y (Geoff mid-split) → exclude the pair.
        base = re.sub(r"[a-z]$", "", mid)
        if mid != base and base in y_ids:
            warn(mid, f"both {base} and {mid} are Y (mid-split) — pair EXCLUDED")
            continue
        # Exact base + one letter — startswith would false-match P1.M10 as a
        # sub-ID of P1.M1.
        sub_siblings = [i for i in y_ids if re.match(re.escape(mid) + r"[a-z]$", i)]
        if sub_siblings:
            warn(mid, f"both {mid} and {sub_siblings} are Y (mid-split) — pair EXCLUDED")
            continue

        measure = build_measure(row, dim_entry, pillar_names, existing_map)
        charted.append(measure)
        gaps[mid] = measure_gaps(row, measure)

    # Previously-charted measures that fell out of the Y set → drop w/ warning.
    charted_ids = {m["measureId"] for m in charted}
    for mid in existing_map:
        if mid not in charted_ids:
            warn(mid, "was in pillar-measures.json but no longer charted "
                      "(un-finalized or newly excluded) — DROPPED")

    charted.sort(key=lambda m: sort_key(m, dim_lookup))

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(charted, f, indent=2, ensure_ascii=False)
        f.write("\n")

    write_gaps_report(gaps, WARNINGS, dim_reconciliation(dim_lookup, id_census))

    new_ids = charted_ids - set(existing_map)
    print(f"\n  Output: {OUTPUT_JSON}")
    print(f"  {len(charted)} measures written "
          f"({len(new_ids)} new, {len(charted_ids & set(existing_map))} refreshed) | "
          f"{len(skipped_bin)} BiN rows skipped (best-in-nation page only)")
    if skipped_bin:
        print(f"    BiN skipped: {', '.join(sorted(skipped_bin))}")

    if WARNINGS:
        print(f"\n  === WARNING SUMMARY ({len(WARNINGS)}) — per-wave to-do list ===")
        for mid, msg in WARNINGS:
            print(f"    [{mid}] {msg}")
    else:
        print("\n  No warnings.")
    print("  Done!")


if __name__ == "__main__":
    main()
